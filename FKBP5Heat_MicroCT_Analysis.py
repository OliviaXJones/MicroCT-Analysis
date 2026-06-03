import sys
import os
import csv
import json
from pathlib import Path

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog, QTextEdit,
    QProgressBar, QMessageBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QDialog, QComboBox, QDialogButtonBox,
)
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtGui import QFont

import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

LINEAGES = ['Pre-heat', 'F0', 'F1', 'F2']

STUDY_PRESETS = {
    'FKBP5 Heat': {
        'W':  'Wildtype', 'H':  'Heterozygous', 'M':  'Mutant',
        '2W': 'Wildtype', '2H': 'Heterozygous', '2M': 'Mutant',
    },
    'FKBP5 New': {
        'W': 'Wildtype',
        'Z': 'Heterozygous',
        'X': 'Mutant',
    },
}

LINEAGE_SUFFIX = {'Pre-heat': '', 'F0': ' F0', 'F1': ' F1', 'F2': ' F2'}

SHEETS = [
    'Wildtype', 'Wildtype F0', 'Wildtype F1', 'Wildtype F2',
    'Mutant', 'Mutant F0', 'Mutant F1', 'Mutant F2',
    'Heterozygous', 'Heterozygous F0', 'Heterozygous F1', 'Heterozygous F2',
]

CORT_HEADERS = [
    'Mouse Code',
    'Total VOI Volume (TV) [mm³]',
    'Object Volume (Obj.V) [mm³]',
    'Structure Thickness (St.Th) [mm]',
    'Medullary Volume (Med.V) [mm³]',
    'vTMD',
]

TRAB_HEADERS = [
    'Mouse Code',
    'Percent Bone Volume (BV/TV) [%]',
    'Bone Surface/Volume Ratio (BS/BV) [1/mm]',
    'Trabecular Pattern Factor (Tb.Pf) [1/mm]',
    'Trabecular Thickness (Tb.Th) [mm]',
    'Trabecular Number (Tb.N) [1/mm]',
    'Trabecular Separation (Tb.Sp) [mm]',
    'Connectivity Density (Conn.Dn) [1/mm³]',
    'vBMD',
]

# Parameters for CSV generation — (header_key, filename_abbreviation)
CORT_PARAMS = [
    ('Total VOI Volume (TV) [mm³]',    'TV'),
    ('Object Volume (Obj.V) [mm³]',    'Obj.V'),
    ('Structure Thickness (St.Th) [mm]', 'St.Th'),
    ('Medullary Volume (Med.V) [mm³]', 'Med.V'),
    ('vTMD',                            'vTMD'),
]

TRAB_PARAMS = [
    ('Percent Bone Volume (BV/TV) [%]',          'BV-TV'),
    ('Bone Surface/Volume Ratio (BS/BV) [1/mm]', 'BS-BV'),
    ('Trabecular Pattern Factor (Tb.Pf) [1/mm]', 'Tb.Pf'),
    ('Trabecular Thickness (Tb.Th) [mm]',        'Tb.Th'),
    ('Trabecular Number (Tb.N) [1/mm]',          'Tb.N'),
    ('Trabecular Separation (Tb.Sp) [mm]',       'Tb.Sp'),
    ('Connectivity Density (Conn.Dn) [1/mm³]',  'Conn.Dn'),
    ('vBMD',                                     'vBMD'),
]

GENOTYPE_ORDER = ['Wildtype', 'Mutant', 'Heterozygous']

SEX_BONE_FOLDERS = {
    ('M', 'cortical'):   'Male Cortical',
    ('M', 'trabecular'): 'Male Trabecular',
    ('F', 'cortical'):   'Female Cortical',
    ('F', 'trabecular'): 'Female Trabecular',
}


# ---------------------------------------------------------------------------
# Mouse code parsing
# ---------------------------------------------------------------------------

def parse_mouse_code(code, genotype_map):
    """
    Parse 'W.12.M21' or 'Z.4.F3' into (genotype_full, age_str, sex).
    Returns None if the code doesn't match the expected format.
    """
    parts = code.split('.')
    if len(parts) != 3:
        return None
    geno_code, age_str, sex_id = parts
    geno_full = genotype_map.get(geno_code.upper())
    if not geno_full:
        return None
    sex = sex_id[0].upper() if sex_id else None
    if sex not in ('M', 'F'):
        return None
    return geno_full, age_str, sex


def sheet_name_for(geno_full, lineage):
    return geno_full + LINEAGE_SUFFIX.get(lineage, '')


# ---------------------------------------------------------------------------
# File parsing helpers
# ---------------------------------------------------------------------------

def mouse_code_from_path(filepath):
    return Path(filepath).stem.split('_')[0]


def extract_field(filepath, prefix):
    """Return float from a CSV line starting with prefix; value is at index 2."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                if line.startswith(prefix):
                    parts = line.strip().split(',')
                    return float(parts[2]) if len(parts) > 2 else None
    except Exception:
        pass
    return None


def extract_hist_mean(filepath, key):
    """Return float from hist file where line starts with key; value is at index 1."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                if line.startswith(key):
                    parts = line.strip().split(',')
                    return float(parts[1]) if len(parts) > 1 else None
    except Exception:
        pass
    return None


def find_sibling_hist(directory, pattern):
    try:
        for f in os.listdir(directory):
            if pattern in f.lower() and f.lower().endswith('.txt'):
                return os.path.join(directory, f)
    except Exception:
        pass
    return None


def find_data_files(root_dir):
    cort_files, trab_files = [], []
    for dirpath, _, filenames in os.walk(root_dir):
        for f in filenames:
            lower = f.lower()
            if '_rec_tra_voi_' in lower and lower.endswith('.txt'):
                path = os.path.join(dirpath, f)
                if lower.endswith('3dcort.txt') or lower.endswith('3d_cort.txt'):
                    cort_files.append(path)
                elif lower.endswith('3dtrab.txt') or lower.endswith('3d_trab.txt'):
                    trab_files.append(path)
    return cort_files, trab_files


def process_cortical_file(filepath):
    tv    = extract_field(filepath, 'Total VOI volume,TV,')
    obj_v = extract_field(filepath, 'Object volume,Obj.V,')
    st_th = extract_field(filepath, 'Structure thickness,St.Th,')
    missing = [name for name, val in [('TV', tv), ('Obj.V', obj_v), ('St.Th', st_th)] if val is None]
    if missing:
        raise ValueError(f"missing fields {missing} in {os.path.basename(filepath)}")
    med_v = tv - obj_v
    vbmd  = None
    hist_path = find_sibling_hist(os.path.dirname(filepath), 'histcort')
    if hist_path:
        vbmd = extract_hist_mean(hist_path, 'Mean:')
    return {
        'Mouse Code':                          mouse_code_from_path(filepath),
        'Total VOI Volume (TV) [mm³]':    tv,
        'Object Volume (Obj.V) [mm³]':    obj_v,
        'Structure Thickness (St.Th) [mm]':    st_th,
        'Medullary Volume (Med.V) [mm³]': med_v,
        'vTMD':                                vbmd,
    }


def process_trabecular_file(filepath):
    vbmd = None
    hist_path = find_sibling_hist(os.path.dirname(filepath), 'histtrab')
    if hist_path:
        vbmd = extract_hist_mean(hist_path, 'Mean (total):')
    bvtv   = extract_field(filepath, 'Percent bone volume,BV/TV,')
    bsbv   = extract_field(filepath, 'Bone surface / volume ratio,BS/BV,')
    tbpf   = extract_field(filepath, 'Trabecular pattern factor,Tb.Pf,')
    tbth   = extract_field(filepath, 'Trabecular thickness,Tb.Th,')
    tbn    = extract_field(filepath, 'Trabecular number,Tb.N,')
    tbsp   = extract_field(filepath, 'Trabecular separation,Tb.Sp,')
    conndn = extract_field(filepath, 'Connectivity density,Conn.Dn,')
    missing = [name for name, val in [
        ('BV/TV', bvtv), ('BS/BV', bsbv), ('Tb.Pf', tbpf), ('Tb.Th', tbth),
        ('Tb.N', tbn), ('Tb.Sp', tbsp), ('Conn.Dn', conndn),
    ] if val is None]
    if missing:
        raise ValueError(f"missing fields {missing} in {os.path.basename(filepath)}")
    return {
        'Mouse Code':                                   mouse_code_from_path(filepath),
        'Percent Bone Volume (BV/TV) [%]':              bvtv,
        'Bone Surface/Volume Ratio (BS/BV) [1/mm]':     bsbv,
        'Trabecular Pattern Factor (Tb.Pf) [1/mm]':     tbpf,
        'Trabecular Thickness (Tb.Th) [mm]':            tbth,
        'Trabecular Number (Tb.N) [1/mm]':              tbn,
        'Trabecular Separation (Tb.Sp) [mm]':           tbsp,
        'Connectivity Density (Conn.Dn) [1/mm³]':       conndn,
        'vBMD':                                         vbmd,
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

_HEADER_FONT = XLFont(bold=True)
_HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')


def _write_sheet(ws, headers, rows):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row_idx, row in enumerate(sorted(rows, key=lambda r: r.get('Mouse Code', '')), 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=row.get(h))
    for col_cells in ws.columns:
        width = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 40)


def _make_workbook():
    wb = openpyxl.Workbook()
    wb.active.title = SHEETS[0]
    for s in SHEETS[1:]:
        wb.create_sheet(s)
    return wb


def write_all_xlsx(output_dir, data):
    """
    data: {(sex, bone_type): {sheet_name: [row_dict, ...]}}
    Writes four .xlsx files into output_dir.
    """
    file_map = {
        ('M', 'cortical'):   ('Male_Cortical.xlsx',    CORT_HEADERS),
        ('M', 'trabecular'): ('Male_Trabecular.xlsx',   TRAB_HEADERS),
        ('F', 'cortical'):   ('Female_Cortical.xlsx',   CORT_HEADERS),
        ('F', 'trabecular'): ('Female_Trabecular.xlsx', TRAB_HEADERS),
    }
    for (sex, bone_type), (filename, headers) in file_map.items():
        wb = _make_workbook()
        sheet_data = data.get((sex, bone_type), {})
        for sheet_name in SHEETS:
            _write_sheet(wb[sheet_name], headers, sheet_data.get(sheet_name, []))
        wb.save(os.path.join(output_dir, filename))


# ---------------------------------------------------------------------------
# CSV analysis output
# ---------------------------------------------------------------------------

def _write_csv(filepath, columns, groups):
    """Write a columnar CSV; groups is {col_name: [values]}. Skips file if all empty."""
    max_len = max((len(groups.get(c, [])) for c in columns), default=0)
    if max_len == 0:
        return
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        for i in range(max_len):
            writer.writerow([
                groups.get(col, [])[i] if i < len(groups.get(col, [])) else ''
                for col in columns
            ])


def generate_csvs(output_dir, all_rows):
    """
    Generate Genotype Analysis and Lineage Analysis CSV folders.
    all_rows: list of row dicts with metadata keys _sex, _bone_type, _geno_full, _age, _lineage.
    """
    ages = sorted(
        set(r['_age'] for r in all_rows),
        key=lambda x: int(x) if x.isdigit() else 0
    )

    analyses = [
        ('Genotype Analysis', False, GENOTYPE_ORDER),
        ('Lineage Analysis',  True,  SHEETS),
    ]

    for analysis_name, use_all_lineages, columns in analyses:
        for (sex, bone_type), folder_name in SEX_BONE_FOLDERS.items():
            params = CORT_PARAMS if bone_type == 'cortical' else TRAB_PARAMS
            folder_path = os.path.join(output_dir, analysis_name, folder_name)

            sex_bone_rows = [
                r for r in all_rows
                if r['_sex'] == sex and r['_bone_type'] == bone_type
            ]

            for age in ages:
                age_rows = [r for r in sex_bone_rows if r['_age'] == age]
                if not use_all_lineages:
                    # Genotype analysis: pre-heat only
                    age_rows = [r for r in age_rows if r['_lineage'] == 'Pre-heat']
                if not age_rows:
                    continue

                for param_key, param_abbrev in params:
                    groups = {}
                    for row in age_rows:
                        if use_all_lineages:
                            col = sheet_name_for(row['_geno_full'], row['_lineage'])
                        else:
                            col = row['_geno_full']
                        val = row.get(param_key)
                        if val is not None:
                            groups.setdefault(col, []).append(val)

                    filepath = os.path.join(folder_path, f'{age}wk_{param_abbrev}.csv')
                    _write_csv(filepath, columns, groups)


# ---------------------------------------------------------------------------
# Worker thread
# ---------------------------------------------------------------------------

class ProcessWorker(QThread):
    log      = pyqtSignal(str)
    progress = pyqtSignal(int)
    finished = pyqtSignal(bool, str)

    def __init__(self, folder_entries, output_dir, genotype_map):
        super().__init__()
        self.folder_entries = folder_entries  # [(folder_path, lineage), ...]
        self.output_dir = output_dir
        self.genotype_map = genotype_map

    def run(self):
        try:
            all_cort, all_trab = [], []
            for folder, lineage in self.folder_entries:
                self.log.emit(f'Scanning [{lineage}]: {folder}')
                c, t = find_data_files(folder)
                all_cort.extend((f, lineage) for f in c)
                all_trab.extend((f, lineage) for f in t)

            total = len(all_cort) + len(all_trab)
            if total == 0:
                self.finished.emit(False, 'No matching files found in any selected folder.')
                return

            self.log.emit(
                f'Found {len(all_cort)} cortical and {len(all_trab)} trabecular file(s) total.'
            )

            xlsx_data = {}   # {(sex, bone_type): {sheet: [row, ...]}}
            all_rows  = []   # enriched rows for CSV generation
            skipped   = []
            done      = 0

            for filepath, lineage in all_cort:
                self.log.emit(f'  [cortical]   {os.path.basename(filepath)}')
                try:
                    row = process_cortical_file(filepath)
                    parsed = parse_mouse_code(row['Mouse Code'], self.genotype_map)
                    if parsed is None:
                        self.log.emit(f'    WARNING: unrecognised code "{row["Mouse Code"]}" — skipped.')
                        skipped.append(row['Mouse Code'])
                    else:
                        geno_full, age, sex = parsed
                        row.update(_sex=sex, _geno_full=geno_full, _age=age,
                                   _lineage=lineage, _bone_type='cortical')
                        all_rows.append(row)
                        sheet = sheet_name_for(geno_full, lineage)
                        xlsx_data.setdefault((sex, 'cortical'), {}).setdefault(sheet, []).append(row)
                except ValueError as e:
                    self.log.emit(f'    ERROR: {e} — skipped.')
                    skipped.append(os.path.basename(filepath))
                done += 1
                self.progress.emit(int(done / total * 50))  # first half = file processing

            for filepath, lineage in all_trab:
                self.log.emit(f'  [trabecular] {os.path.basename(filepath)}')
                try:
                    row = process_trabecular_file(filepath)
                    parsed = parse_mouse_code(row['Mouse Code'], self.genotype_map)
                    if parsed is None:
                        self.log.emit(f'    WARNING: unrecognised code "{row["Mouse Code"]}" — skipped.')
                        skipped.append(row['Mouse Code'])
                    else:
                        geno_full, age, sex = parsed
                        row.update(_sex=sex, _geno_full=geno_full, _age=age,
                                   _lineage=lineage, _bone_type='trabecular')
                        all_rows.append(row)
                        sheet = sheet_name_for(geno_full, lineage)
                        xlsx_data.setdefault((sex, 'trabecular'), {}).setdefault(sheet, []).append(row)
                except ValueError as e:
                    self.log.emit(f'    ERROR: {e} — skipped.')
                    skipped.append(os.path.basename(filepath))
                done += 1
                self.progress.emit(int(done / total * 50))

            self.log.emit('Writing Excel files…')
            write_all_xlsx(self.output_dir, xlsx_data)
            self.progress.emit(75)

            self.log.emit('Generating analysis CSVs…')
            generate_csvs(self.output_dir, all_rows)
            self.progress.emit(100)

            msg = (
                f'Done!\n'
                f'  • 4 master .xlsx files → {self.output_dir}\n'
                f'  • Genotype Analysis CSVs → {os.path.join(self.output_dir, "Genotype Analysis")}\n'
                f'  • Lineage Analysis CSVs  → {os.path.join(self.output_dir, "Lineage Analysis")}'
            )
            if skipped:
                msg += f'\n\nSkipped {len(skipped)} unrecognised code(s): {", ".join(skipped)}'
            self.finished.emit(True, msg)

        except Exception as e:
            import traceback
            self.finished.emit(False, f'Error: {e}\n{traceback.format_exc()}')


# ---------------------------------------------------------------------------
# Lineage picker dialog
# ---------------------------------------------------------------------------

class LineageDialog(QDialog):
    def __init__(self, folder_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Select Lineage')
        self.setMinimumWidth(440)
        layout = QVBoxLayout(self)
        lbl = QLabel(f'Folder:\n{folder_path}')
        lbl.setWordWrap(True)
        layout.addWidget(lbl)
        layout.addWidget(QLabel('\nWhich lineage does this folder represent?'))
        self.combo = QComboBox()
        self.combo.addItems(LINEAGES)
        layout.addWidget(self.combo)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def lineage(self):
        return self.combo.currentText()


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self._current_study = None
        self.setWindowTitle('MicroCT Data Extractor — FKBP5')
        self.setMinimumWidth(780)
        self._build_ui()
        self._current_study = self.study_selector.currentText()
        self._load_config()

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        study_row = QHBoxLayout()
        study_row.addWidget(QLabel('Study:'))
        self.study_selector = QComboBox()
        self.study_selector.addItems(list(STUDY_PRESETS.keys()))
        self.study_selector.currentTextChanged.connect(self._on_study_changed)
        study_row.addWidget(self.study_selector)
        study_row.addStretch()
        layout.addLayout(study_row)

        layout.addWidget(QLabel('Source folders (add one per lineage):'))

        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(['Folder', 'Lineage'])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setMinimumHeight(120)
        layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        add_btn = QPushButton('Add Folder…')
        add_btn.clicked.connect(self._add_folder)
        self.remove_btn = QPushButton('Remove Selected')
        self.remove_btn.setEnabled(False)
        self.remove_btn.clicked.connect(self._remove_folder)
        self.table.itemSelectionChanged.connect(
            lambda: self.remove_btn.setEnabled(bool(self.table.selectedItems()))
        )
        btn_row.addWidget(add_btn)
        btn_row.addWidget(self.remove_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        out_row = QHBoxLayout()
        lbl = QLabel('Output folder:')
        lbl.setFixedWidth(100)
        self.output_edit = QLineEdit()
        self.output_edit.setPlaceholderText(
            'Folder where .xlsx files and Analysis CSVs will be saved…'
        )
        self.output_edit.setReadOnly(True)
        out_btn = QPushButton('Browse…')
        out_btn.setFixedWidth(80)
        out_btn.clicked.connect(self._pick_output)
        out_row.addWidget(lbl)
        out_row.addWidget(self.output_edit)
        out_row.addWidget(out_btn)
        layout.addLayout(out_row)

        self.run_btn = QPushButton('Extract Data')
        self.run_btn.setEnabled(False)
        self.run_btn.setFixedHeight(36)
        self.run_btn.clicked.connect(self._run)
        layout.addWidget(self.run_btn)

        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFont(QFont('Courier New', 9))
        self.log_box.setMinimumHeight(180)
        layout.addWidget(self.log_box)

    def _add_folder(self):
        path = QFileDialog.getExistingDirectory(self, 'Select Data Folder')
        if not path:
            return
        dlg = LineageDialog(path, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem(path))
        self.table.setItem(row, 1, QTableWidgetItem(dlg.lineage()))
        self._refresh_run()
        self._save_config()

    def _remove_folder(self):
        for r in sorted({i.row() for i in self.table.selectedItems()}, reverse=True):
            self.table.removeRow(r)
        self._refresh_run()
        self._save_config()

    def _pick_output(self):
        path = QFileDialog.getExistingDirectory(self, 'Select Output Folder')
        if path:
            self.output_edit.setText(path)
            self._refresh_run()
            self._save_config()

    def _refresh_run(self):
        self.run_btn.setEnabled(
            self.table.rowCount() > 0 and bool(self.output_edit.text())
        )

    def _run(self):
        self.run_btn.setEnabled(False)
        self.log_box.clear()
        self.progress_bar.setValue(0)
        entries = [
            (self.table.item(r, 0).text(), self.table.item(r, 1).text())
            for r in range(self.table.rowCount())
        ]
        genotype_map = STUDY_PRESETS[self.study_selector.currentText()]
        self._worker = ProcessWorker(entries, self.output_edit.text(), genotype_map)
        self._worker.log.connect(self.log_box.append)
        self._worker.progress.connect(self.progress_bar.setValue)
        self._worker.finished.connect(self._on_done)
        self._worker.start()

    def _config_path(self):
        base = os.path.dirname(os.path.abspath(sys.argv[0]))
        return os.path.join(base, 'fkbp5_paths.json')

    def _load_config(self):
        study = self.study_selector.currentText()
        try:
            with open(self._config_path(), 'r') as f:
                config = json.load(f)
            data = config.get(study, {})
            folders = data.get('folders', [])
            output = data.get('output', '')
        except (FileNotFoundError, json.JSONDecodeError):
            folders = []
            output = ''
        self.table.setRowCount(0)
        for path, lineage in folders:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(path))
            self.table.setItem(row, 1, QTableWidgetItem(lineage))
        self.output_edit.setText(output)
        self._refresh_run()

    def _save_config(self, study=None):
        if study is None:
            study = self.study_selector.currentText()
        folders = [
            (self.table.item(r, 0).text(), self.table.item(r, 1).text())
            for r in range(self.table.rowCount())
        ]
        output = self.output_edit.text()
        config = {}
        try:
            with open(self._config_path(), 'r') as f:
                config = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            pass
        config[study] = {'folders': folders, 'output': output}
        try:
            with open(self._config_path(), 'w') as f:
                json.dump(config, f, indent=2)
        except Exception:
            pass

    def _on_study_changed(self, new_study):
        if self._current_study and self._current_study != new_study:
            self._save_config(self._current_study)
        self._current_study = new_study
        self._load_config()

    def _on_done(self, ok, msg):
        self.log_box.append(msg)
        self.run_btn.setEnabled(True)
        if ok:
            QMessageBox.information(self, 'Complete', msg)
        else:
            QMessageBox.warning(self, 'Error', msg)


# ---------------------------------------------------------------------------

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
