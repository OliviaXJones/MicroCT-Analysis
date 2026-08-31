import sys
import os
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox,
    QFileDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QDialog, QTextEdit, QMessageBox, QDialogButtonBox, QInputDialog,
    QAbstractItemView, QProgressBar,
)
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtGui import QFont

def _app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(_app_dir(), "studies.json")


# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_config(config):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=2)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CORT_HEADERS = [
    "Mouse Code",
    "Total VOI Volume (TV) [mm³]",
    "Object Volume (Obj.V) [mm³]",
    "Structure Thickness (St.Th) [mm]",
    "Medullary Volume (Med.V) [mm³]",
    "vTMD",
]

TRAB_HEADERS = [
    "Mouse Code",
    "Percent Bone Volume (BV/TV) [%]",
    "Bone Surface/Volume Ratio (BS/BV) [1/mm]",
    "Trabecular Pattern Factor (Tb.Pf) [1/mm]",
    "Trabecular Thickness (Tb.Th) [mm]",
    "Trabecular Number (Tb.N) [1/mm]",
    "Trabecular Separation (Tb.Sp) [mm]",
    "Connectivity Density (Conn.Dn) [1/mm³]",
    "vBMD",
]

FEMUR_SEX_FILE_MAP = {
    ("M", "cortical"):   ("Male_Cortical.xlsx",    CORT_HEADERS),
    ("M", "trabecular"): ("Male_Trabecular.xlsx",  TRAB_HEADERS),
    ("F", "cortical"):   ("Female_Cortical.xlsx",  CORT_HEADERS),
    ("F", "trabecular"): ("Female_Trabecular.xlsx", TRAB_HEADERS),
}

SPINE_SEX_FILE_MAP = {
    "M": ("Male_Spine.xlsx",   TRAB_HEADERS),
    "F": ("Female_Spine.xlsx", TRAB_HEADERS),
}


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------

def mouse_code_from_path(filepath):
    return Path(filepath).stem.split("_")[0]


def extract_field(filepath, prefix):
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.startswith(prefix):
                    parts = line.strip().split(",")
                    return float(parts[2]) if len(parts) > 2 else None
    except Exception:
        pass
    return None


def extract_hist_mean(filepath, key):
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.startswith(key):
                    parts = line.strip().split(",")
                    return float(parts[1]) if len(parts) > 1 else None
    except Exception:
        pass
    return None


def find_sibling_hist(directory, pattern):
    try:
        for f in os.listdir(directory):
            if pattern in f.lower() and f.lower().endswith(".txt"):
                return os.path.join(directory, f)
    except Exception:
        pass
    return None


def find_data_files(root_dir):
    cort_files, trab_files = [], []
    for dirpath, _, filenames in os.walk(root_dir):
        for f in filenames:
            lower = f.lower()
            if "_rec_tra_voi_" in lower and lower.endswith(".txt"):
                path = os.path.join(dirpath, f)
                if lower.endswith("3dcort.txt") or lower.endswith("3d_cort.txt"):
                    cort_files.append(path)
                elif lower.endswith("3dtrab.txt") or lower.endswith("3d_trab.txt"):
                    trab_files.append(path)
    return cort_files, trab_files


def process_cortical_file(filepath):
    tv    = extract_field(filepath, "Total VOI volume,TV,")
    obj_v = extract_field(filepath, "Object volume,Obj.V,")
    st_th = extract_field(filepath, "Structure thickness,St.Th,")
    missing = [name for name, val in [("TV", tv), ("Obj.V", obj_v), ("St.Th", st_th)] if val is None]
    if missing:
        raise ValueError(f"missing fields {missing} in {os.path.basename(filepath)}")
    med_v = tv - obj_v
    vbmd  = None
    hist_path = find_sibling_hist(os.path.dirname(filepath), "histcort")
    if hist_path:
        vbmd = extract_hist_mean(hist_path, "Mean:")
    return {
        "Mouse Code":                       mouse_code_from_path(filepath),
        "Total VOI Volume (TV) [mm³]":      tv,
        "Object Volume (Obj.V) [mm³]":      obj_v,
        "Structure Thickness (St.Th) [mm]": st_th,
        "Medullary Volume (Med.V) [mm³]":   med_v,
        "vTMD":                             vbmd,
    }


def process_trabecular_file(filepath):
    vbmd = None
    hist_path = find_sibling_hist(os.path.dirname(filepath), "histtrab")
    if hist_path:
        vbmd = extract_hist_mean(hist_path, "Mean (total):")
    bvtv   = extract_field(filepath, "Percent bone volume,BV/TV,")
    bsbv   = extract_field(filepath, "Bone surface / volume ratio,BS/BV,")
    tbpf   = extract_field(filepath, "Trabecular pattern factor,Tb.Pf,")
    tbth   = extract_field(filepath, "Trabecular thickness,Tb.Th,")
    tbn    = extract_field(filepath, "Trabecular number,Tb.N,")
    tbsp   = extract_field(filepath, "Trabecular separation,Tb.Sp,")
    conndn = extract_field(filepath, "Connectivity density,Conn.Dn,")
    missing = [name for name, val in [
        ("BV/TV", bvtv), ("BS/BV", bsbv), ("Tb.Pf", tbpf), ("Tb.Th", tbth),
        ("Tb.N", tbn), ("Tb.Sp", tbsp), ("Conn.Dn", conndn),
    ] if val is None]
    if missing:
        raise ValueError(f"missing fields {missing} in {os.path.basename(filepath)}")
    return {
        "Mouse Code":                                   mouse_code_from_path(filepath),
        "Percent Bone Volume (BV/TV) [%]":              bvtv,
        "Bone Surface/Volume Ratio (BS/BV) [1/mm]":     bsbv,
        "Trabecular Pattern Factor (Tb.Pf) [1/mm]":     tbpf,
        "Trabecular Thickness (Tb.Th) [mm]":            tbth,
        "Trabecular Number (Tb.N) [1/mm]":              tbn,
        "Trabecular Separation (Tb.Sp) [mm]":           tbsp,
        "Connectivity Density (Conn.Dn) [1/mm³]":       conndn,
        "vBMD":                                         vbmd,
    }


# ---------------------------------------------------------------------------
# Mouse code parsing
# ---------------------------------------------------------------------------

# Matches codes like CC1M, CC1, MC2F, HC10, HCF10, HCM10 — prefix (letters),
# ID (digits), optional trailing sex marker (M or F).
_CODE_RE = re.compile(r'^([A-Za-z]+)(\d+)([MF]?)$', re.IGNORECASE)


def parse_mouse_code(code, group_map, study_sex):
    """
    Parses PREFIX+ID[+SEX] or PREFIX[SEX]+ID without any separator.
    Sex can appear after the ID (HC10F), or embedded at the end of the
    prefix before the ID (HCF10 / HCM10).
    """
    m = _CODE_RE.match(code)
    if not m:
        return None
    prefix   = m.group(1).upper()
    sex_char = m.group(3).upper()

    lookup_prefix = prefix

    if study_sex == "Mixed":
        if sex_char:
            sex = sex_char
        elif prefix.endswith(("F", "M")):
            sex = prefix[-1]
            lookup_prefix = prefix[:-1]
        else:
            return None
    else:
        sex = "M" if study_sex == "Male" else "F"

    group_name = group_map.get(lookup_prefix)
    if not group_name:
        return None
    return group_name, sex


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

_HEADER_FONT = XLFont(bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _write_sheet(ws, headers, rows):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row_idx, row in enumerate(sorted(rows, key=lambda r: r.get("Mouse Code", "")), 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=row.get(h))
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 40)


def write_femur_xlsx(output_dir, data, group_names, study_sex):
    sexes = {"Male": ["M"], "Female": ["F"], "Mixed": ["M", "F"]}[study_sex]
    sheets = group_names if group_names else ["Data"]
    for (sex, bone_type), (filename, headers) in FEMUR_SEX_FILE_MAP.items():
        if sex not in sexes:
            continue
        wb = openpyxl.Workbook()
        wb.active.title = sheets[0]
        for name in sheets[1:]:
            wb.create_sheet(name)
        sheet_data = data.get((sex, bone_type), {})
        for sheet_name in sheets:
            _write_sheet(wb[sheet_name], headers, sheet_data.get(sheet_name, []))
        wb.save(os.path.join(output_dir, filename))


def write_spine_xlsx(output_dir, data, group_names, study_sex):
    sexes = {"Male": ["M"], "Female": ["F"], "Mixed": ["M", "F"]}[study_sex]
    sheets = group_names if group_names else ["Data"]
    for sex in sexes:
        filename, headers = SPINE_SEX_FILE_MAP[sex]
        wb = openpyxl.Workbook()
        wb.active.title = sheets[0]
        for name in sheets[1:]:
            wb.create_sheet(name)
        sheet_data = data.get((sex, "trabecular"), {})
        for sheet_name in sheets:
            _write_sheet(wb[sheet_name], headers, sheet_data.get(sheet_name, []))
        wb.save(os.path.join(output_dir, filename))


# ---------------------------------------------------------------------------
# Worker thread
# ---------------------------------------------------------------------------

class ProcessWorker(QThread):
    log      = pyqtSignal(str)
    progress = pyqtSignal(int)
    finished = pyqtSignal(bool, str)

    def __init__(self, data_folder, output_dir, study_cfg, bone_type):
        super().__init__()
        self.data_folder = data_folder
        self.output_dir  = output_dir
        self.study_cfg   = study_cfg
        self.bone_type   = bone_type  # "femur" or "spine"

    def run(self):
        try:
            cfg         = self.study_cfg
            study_sex   = cfg.get("sex", "Mixed")
            group_map   = {g["prefix"].upper(): g["group_name"] for g in cfg.get("group_map", [])}
            group_order = [g["group_name"] for g in cfg.get("group_map", [])]

            self.log.emit(f"Scanning: {self.data_folder}")
            cort_files, trab_files = find_data_files(self.data_folder)

            if self.bone_type == "spine":
                cort_files = []  # spine analysis uses trabecular only

            total = len(cort_files) + len(trab_files)
            if total == 0:
                self.finished.emit(False, "No matching CT files found in the selected folder.")
                return

            if self.bone_type == "spine":
                self.log.emit(f"Found {len(trab_files)} spine trabecular file(s).")
            else:
                self.log.emit(f"Found {len(cort_files)} cortical and {len(trab_files)} trabecular file(s).")

            xlsx_data = {}
            skipped   = []
            done      = 0

            for filepath in cort_files:
                self.log.emit(f"  [cortical]   {os.path.basename(filepath)}")
                try:
                    row    = process_cortical_file(filepath)
                    parsed = parse_mouse_code(row["Mouse Code"], group_map, study_sex)
                    if parsed is None:
                        self.log.emit(f'    WARNING: unrecognised code "{row["Mouse Code"]}" — skipped.')
                        skipped.append(row["Mouse Code"])
                    else:
                        group_name, sex = parsed
                        xlsx_data.setdefault((sex, "cortical"), {}).setdefault(group_name, []).append(row)
                except ValueError as e:
                    self.log.emit(f"    ERROR: {e} — skipped.")
                    skipped.append(os.path.basename(filepath))
                done += 1
                self.progress.emit(int(done / total * 80))

            trab_label = "spine" if self.bone_type == "spine" else "trabecular"
            for filepath in trab_files:
                self.log.emit(f"  [{trab_label}]  {os.path.basename(filepath)}")
                try:
                    row    = process_trabecular_file(filepath)
                    parsed = parse_mouse_code(row["Mouse Code"], group_map, study_sex)
                    if parsed is None:
                        self.log.emit(f'    WARNING: unrecognised code "{row["Mouse Code"]}" — skipped.')
                        skipped.append(row["Mouse Code"])
                    else:
                        group_name, sex = parsed
                        xlsx_data.setdefault((sex, "trabecular"), {}).setdefault(group_name, []).append(row)
                except ValueError as e:
                    self.log.emit(f"    ERROR: {e} — skipped.")
                    skipped.append(os.path.basename(filepath))
                done += 1
                self.progress.emit(int(done / total * 80))

            self.log.emit("Writing Excel files…")
            if self.bone_type == "spine":
                write_spine_xlsx(self.output_dir, xlsx_data, group_order, study_sex)
            else:
                write_femur_xlsx(self.output_dir, xlsx_data, group_order, study_sex)
            self.progress.emit(100)

            msg = f"Done!\n  • Output → {self.output_dir}"
            if skipped:
                msg += f"\n\nSkipped {len(skipped)} unrecognised code(s): {', '.join(skipped)}"
            self.finished.emit(True, msg)

        except Exception as e:
            import traceback
            self.finished.emit(False, f"Error: {e}\n{traceback.format_exc()}")


# ---------------------------------------------------------------------------
# Group map editor widget
# ---------------------------------------------------------------------------

class GroupMapEditor(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Prefix", "Group Name"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setMinimumHeight(120)
        layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("Add Row")
        add_btn.clicked.connect(self._add_row)
        self.remove_btn = QPushButton("Remove Selected")
        self.remove_btn.setEnabled(False)
        self.remove_btn.clicked.connect(self._remove_row)
        self.table.itemSelectionChanged.connect(
            lambda: self.remove_btn.setEnabled(bool(self.table.selectedItems()))
        )
        btn_row.addWidget(add_btn)
        btn_row.addWidget(self.remove_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

    def _add_row(self):
        r = self.table.rowCount()
        self.table.insertRow(r)
        self.table.setItem(r, 0, QTableWidgetItem(""))
        self.table.setItem(r, 1, QTableWidgetItem(""))

    def _remove_row(self):
        for r in sorted({i.row() for i in self.table.selectedItems()}, reverse=True):
            self.table.removeRow(r)

    def get_groups(self):
        groups = []
        for r in range(self.table.rowCount()):
            prefix = (self.table.item(r, 0) or QTableWidgetItem("")).text().strip()
            name   = (self.table.item(r, 1) or QTableWidgetItem("")).text().strip()
            if prefix:
                groups.append({"prefix": prefix, "group_name": name})
        return groups

    def set_groups(self, groups):
        self.table.setRowCount(0)
        for g in groups:
            r = self.table.rowCount()
            self.table.insertRow(r)
            self.table.setItem(r, 0, QTableWidgetItem(g.get("prefix", "")))
            self.table.setItem(r, 1, QTableWidgetItem(g.get("group_name", "")))


# ---------------------------------------------------------------------------
# Get Info dialog
# ---------------------------------------------------------------------------

class GetInfoDialog(QDialog):
    def __init__(self, study_name, data, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Study Info — {study_name}")
        self.setMinimumWidth(520)
        layout = QVBoxLayout(self)

        text = QTextEdit()
        text.setReadOnly(True)
        text.setFont(QFont("Courier New", 9))

        sex = data.get("sex", "—")
        fmt = {
            "Male":   "CC1  (prefix + ID)",
            "Female": "CC1  (prefix + ID)",
            "Mixed":  "CC1M / CC1F  (prefix + ID + sex)",
        }.get(sex, "—")

        femur = data.get("femur", {})
        spine = data.get("spine", {})

        # backward compat: old top-level output_folder stored under femur
        if not femur and "output_folder" in data:
            femur = {
                "data_folder":   data.get("data_folder", ""),
                "output_folder": data.get("output_folder", ""),
            }

        lines = [
            f"Study:             {study_name}",
            f"Age:               {data.get('age', '—')}",
            f"Sex:               {sex}",
            f"Mouse Code Format: {fmt}",
            "",
            "Femur:",
            f"  Data Folder:   {femur.get('data_folder', '') or '—'}",
            f"  Output Folder: {femur.get('output_folder', '') or '—'}",
            "",
            "Spine:",
            f"  Data Folder:   {spine.get('data_folder', '') or '—'}",
            f"  Output Folder: {spine.get('output_folder', '') or '—'}",
            "",
            "Group Map:",
        ]
        groups = data.get("group_map", [])
        if groups:
            for g in groups:
                lines.append(f"  {g['prefix']:<10} →  {g.get('group_name', '')}")
        else:
            lines.append("  (none defined)")

        text.setPlainText("\n".join(lines))
        layout.addWidget(text)

        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MicroCT Data Extractor — Normal Study")
        self.setMinimumWidth(720)
        self.config = load_config()
        self._build_ui()
        self._populate_selector()

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        outer = QVBoxLayout(root)
        outer.setSpacing(10)
        outer.setContentsMargins(16, 16, 16, 16)

        # Study selector row
        sel_row = QHBoxLayout()
        sel_row.addWidget(QLabel("Study:"))
        self.study_selector = QComboBox()
        self.study_selector.setMinimumWidth(220)
        self.study_selector.currentTextChanged.connect(self._load_study)
        sel_row.addWidget(self.study_selector)
        new_btn = QPushButton("New Study")
        new_btn.clicked.connect(self._new_study)
        del_btn = QPushButton("Delete Study")
        del_btn.clicked.connect(self._delete_study)
        sel_row.addWidget(new_btn)
        sel_row.addWidget(del_btn)
        sel_row.addStretch()
        outer.addLayout(sel_row)

        # Study config — shared across bones
        cfg_group = QGroupBox("Study Configuration")
        form = QFormLayout(cfg_group)

        self.f_age = QLineEdit("16 Weeks")
        form.addRow(QLabel("Cohort Age:"), self.f_age)

        self.f_sex = QComboBox()
        self.f_sex.addItems(["Male", "Female", "Mixed"])
        form.addRow(QLabel("Cohort Sex:"), self.f_sex)

        outer.addWidget(cfg_group)

        # Group map
        gmap_group = QGroupBox("Group Map  (Prefix → Group Name)")
        gmap_layout = QVBoxLayout(gmap_group)
        gmap_layout.addWidget(QLabel(
            "Map mouse code prefixes to group names.  e.g.  CC → Control,  HC → Heat + Control"
        ))
        self.group_map_editor = GroupMapEditor()
        gmap_layout.addWidget(self.group_map_editor)
        outer.addWidget(gmap_group)

        # Save / Get Info
        study_btns = QHBoxLayout()
        save_btn = QPushButton("Save Study")
        save_btn.clicked.connect(self._save_study)
        info_btn = QPushButton("Get Info")
        info_btn.clicked.connect(self._get_info)
        study_btns.addWidget(save_btn)
        study_btns.addWidget(info_btn)
        study_btns.addStretch()
        outer.addLayout(study_btns)

        # Run section — bone-specific folders
        run_group = QGroupBox("Run")
        run_form = QFormLayout(run_group)

        self.f_bone = QComboBox()
        self.f_bone.addItems(["Femur", "Spine"])
        self.f_bone.currentTextChanged.connect(self._on_bone_changed)
        run_form.addRow(QLabel("Bone:"), self.f_bone)

        data_row = QHBoxLayout()
        self.f_data = QLineEdit()
        self.f_data.setPlaceholderText("Folder containing all nested CT scan data…")
        self.f_data.textChanged.connect(self._refresh_run_btn)
        data_btn = QPushButton("Browse")
        data_btn.clicked.connect(lambda: self._browse_dir(self.f_data))
        data_row.addWidget(self.f_data)
        data_row.addWidget(data_btn)
        run_form.addRow(QLabel("Data Folder:"), data_row)

        out_row = QHBoxLayout()
        self.f_out = QLineEdit()
        self.f_out.setPlaceholderText("Folder where .xlsx files will be saved…")
        self.f_out.textChanged.connect(self._refresh_run_btn)
        out_btn = QPushButton("Browse")
        out_btn.clicked.connect(lambda: self._browse_dir(self.f_out))
        out_row.addWidget(self.f_out)
        out_row.addWidget(out_btn)
        run_form.addRow(QLabel("Output Folder:"), out_row)

        outer.addWidget(run_group)

        self.run_btn = QPushButton("Extract Data")
        self.run_btn.setFixedHeight(36)
        self.run_btn.setEnabled(False)
        self.run_btn.clicked.connect(self._run)
        outer.addWidget(self.run_btn)

        self.progress_bar = QProgressBar()
        outer.addWidget(self.progress_bar)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFont(QFont("Courier New", 9))
        self.log_box.setMinimumHeight(160)
        outer.addWidget(self.log_box)

    # ------------------------------------------------------------------
    # Study selector
    # ------------------------------------------------------------------

    def _populate_selector(self):
        self.study_selector.blockSignals(True)
        self.study_selector.clear()
        self.study_selector.addItems(sorted(self.config.keys()))
        self.study_selector.blockSignals(False)
        if self.study_selector.count():
            self.study_selector.setCurrentIndex(0)
            self._load_study(self.study_selector.currentText())

    def _load_study(self, name):
        if not name:
            return
        d = self.config.get(name, {})
        self.f_age.setText(d.get("age", "16 Weeks"))
        self.f_sex.setCurrentText(d.get("sex", "Male"))
        self.group_map_editor.set_groups(d.get("group_map", []))
        self._load_bone_folders(d)

    def _load_bone_folders(self, d):
        bone = self.f_bone.currentText().lower()
        bone_data = d.get(bone, {})
        # backward compat: old top-level output_folder is treated as femur's
        if not bone_data and bone == "femur" and "output_folder" in d:
            self.f_data.setText(d.get("data_folder", ""))
            self.f_out.setText(d.get("output_folder", ""))
        else:
            self.f_data.setText(bone_data.get("data_folder", ""))
            self.f_out.setText(bone_data.get("output_folder", ""))

    def _on_bone_changed(self, _):
        name = self.study_selector.currentText()
        if name and name in self.config:
            self._load_bone_folders(self.config[name])
        else:
            self.f_data.clear()
            self.f_out.clear()

    def _new_study(self):
        name, ok = QInputDialog.getText(self, "New Study", "Study name:")
        name = name.strip()
        if not ok or not name:
            return
        if name in self.config:
            QMessageBox.warning(self, "Already Exists", f'A study named "{name}" already exists.')
            return
        self.config[name] = {}
        save_config(self.config)
        self.study_selector.blockSignals(True)
        self.study_selector.addItem(name)
        self.study_selector.blockSignals(False)
        self.study_selector.setCurrentText(name)
        self._load_study(name)

    def _delete_study(self):
        name = self.study_selector.currentText()
        if not name:
            return
        reply = QMessageBox.question(
            self, "Delete Study", f'Delete "{name}"?',
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.config.pop(name, None)
        save_config(self.config)
        self.study_selector.removeItem(self.study_selector.currentIndex())

    # ------------------------------------------------------------------
    # Save / Get Info
    # ------------------------------------------------------------------

    def _save_study(self):
        name = self.study_selector.currentText().strip()
        if not name:
            QMessageBox.warning(self, "No Study", "Select or create a study first.")
            return
        bone = self.f_bone.currentText().lower()
        existing = dict(self.config.get(name, {}))
        existing.update({
            "age":       self.f_age.text().strip(),
            "sex":       self.f_sex.currentText(),
            "group_map": self.group_map_editor.get_groups(),
        })
        bone_data = dict(existing.get(bone, {}))
        bone_data.update({
            "data_folder":   self.f_data.text().strip(),
            "output_folder": self.f_out.text().strip(),
        })
        existing[bone] = bone_data
        # remove legacy top-level folder keys on first save in new format
        existing.pop("output_folder", None)
        existing.pop("data_folder", None)
        self.config[name] = existing
        save_config(self.config)
        QMessageBox.information(self, "Saved", f'Study "{name}" saved.')

    def _get_info(self):
        name = self.study_selector.currentText()
        if not name or name not in self.config:
            QMessageBox.warning(self, "Not Saved", "Save the study first before viewing its info.")
            return
        GetInfoDialog(name, self.config[name], self).exec_()

    # ------------------------------------------------------------------
    # Run
    # ------------------------------------------------------------------

    def _refresh_run_btn(self):
        self.run_btn.setEnabled(
            bool(self.f_data.text().strip()) and bool(self.f_out.text().strip())
        )

    def _run(self):
        name = self.study_selector.currentText()
        if not name:
            QMessageBox.warning(self, "No Study", "Select or create a study first.")
            return
        self.run_btn.setEnabled(False)
        self.log_box.clear()
        self.progress_bar.setValue(0)
        bone_type = self.f_bone.currentText().lower()
        study_cfg = dict(self.config.get(name, {}))
        study_cfg.update({
            "sex":       self.f_sex.currentText(),
            "group_map": self.group_map_editor.get_groups(),
        })
        self._worker = ProcessWorker(
            self.f_data.text().strip(),
            self.f_out.text().strip(),
            study_cfg,
            bone_type,
        )
        self._worker.log.connect(self.log_box.append)
        self._worker.progress.connect(self.progress_bar.setValue)
        self._worker.finished.connect(self._on_done)
        self._worker.start()

    def _on_done(self, ok, msg):
        self.log_box.append(msg)
        self.run_btn.setEnabled(True)
        if ok:
            QMessageBox.information(self, "Complete", msg)
        else:
            QMessageBox.warning(self, "Error", msg)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _browse_dir(self, line_edit):
        path = QFileDialog.getExistingDirectory(self, "Select Directory")
        if path:
            line_edit.setText(os.path.normpath(path))


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
